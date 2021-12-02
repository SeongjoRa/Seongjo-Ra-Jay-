from tkinter import *
import tkinter.messagebox as msgbox
import tkinter.font as font
import tkinter.ttk as ttk
from selenium import webdriver
import os
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import requests
from bs4 import BeautifulSoup
from openpyxl import *
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border,Font, PatternFill, Side
from openpyxl.utils.cell import get_column_letter
import re

root = Tk()
root.iconbitmap("window_page_file_document_icon_196480.ico")
root.title("Data Quality Control")
root.geometry("240x320")
root.resizable(False, False)

def Open():
    os.startfile("QA report.xlsx")
def msg1():
    msgbox.showinfo("Trouble Shooting","VPN required for PDP (Search), Product Details, Specifications, WTB/CTA and WTB (New)\n\nSave is a MUST after PDP (Search/Create).\n\nClose ‘QA Report’ before running any applications.\n\nSimply close the 'command window' to stop any running applications. When running again, the application will start from where it stopped. \n\nLocale in 'Input tab' is properly selected? Required for PDP (Search) and Data Comparison (Web/PISA)\n\nCoulumns of '3M ID/SKU' in 'Product Details/Original/PISA' are all in the SAME ORDER? Attributes for your target locale in 'Original' and 'PISA' tabs are all in 'en (English International)'? Required for Data Comparison (Web/PISA)\n\nChrome Browser has been updated? https://chromedriver.chromium.org/")
def msg2():
    msgbox.showinfo("Version","Vesion 1.1")
def msg3():
    msgbox.showinfo("Reort Issue","sra2@mmm.com")

menu = Menu(root)
root.config(menu=menu)

myFont1 = font.Font(family='Franklin Gothic Medium', size=10)
menu_File = Menu(menu, tearoff=0)
menu_File.add_command(label="Open QA report", font=myFont1, command=Open)
menu_File.add_separator()
menu_File.add_command(label="Exit", font=myFont1,  command=root.quit)
menu.add_cascade(label="File", font=myFont1,  menu=menu_File)
menu_Help = Menu(menu, tearoff=0)
menu_Help.add_command(label="Troubleshooting", font=myFont1,  command=msg1)
menu_Help.add_separator()
menu_Help.add_command(label="Report Issue", font=myFont1,  command=msg3)
menu.add_cascade(label="Help", font=myFont1,  menu=menu_Help)
menu_About = Menu(menu, tearoff=0)
menu_About.add_command(label="Version", font=myFont1, command=msg2)
menu.add_cascade(label="About", font=myFont1, menu=menu_About)

p_var = DoubleVar()
pb = ttk.Progressbar(root, maximum=100, variable=p_var)
pb.grid(row=7, column=0, columnspan=2, sticky=N+E+W)

def scrpt1():
    inputs = {'Liveko_KR':'https://www.3m.co.kr/3M/ko_KR/company-kr/search/?Ntt=', 'Prevko_KR':'https://fuzeauth.3m.com/3M/ko_KR/company-kr/search/?Ntt=', 
    'Liveen_PH':'https://www.3mphilippines.com.ph/3M/en_PH/company-ph/search/?Ntt=', 'Preven_PH':'https://fuzeauth.3m.com/3M/en_PH/company-ph/search/?Ntt=',
    'Liveen_AU':'https://www.3m.com.au/3M/en_AU/company-au/search/?Ntt=', 'Preven_AU':'https://fuzeauth.3m.com/3M/en_AU/company-au/search/?Ntt=',
    'Liveen_NZ':'https://www.3mnz.co.nz/3M/en_NZ/company-nz/search/?Ntt=', 'Preven_NZ':'https://fuzeauth.3m.com/3M/en_NZ/company-nz/search/?Ntt=',
    'Liveth_TH':'https://www.3m.co.th/3M/th_TH/company-th/search/?Ntt=', 'Prevth_TH':'https://fuzeauth.3m.com/3M/th_TH/company-th/search/?Ntt=',
    'Liveen_MY':'https://www.3m.com.my/3M/en_MY/company-my/search/?Ntt=', 'Preven_MY':'https://fuzeauth.3m.com/3M/en_MY/company-my/search/?Ntt=',
    'Liveen_SG':'https://www.3m.com.sg/3M/en_SG/company-sg/search/?Ntt=', 'Preven_SG':'https://fuzeauth.3m.com/3M/en_SG/company-sg/search/?Ntt=',
    'Liveen_IN':'https://www.3mindia.in/3M/en_IN/company-in/search/?Ntt=', 'Preven_IN':'https://fuzeauth.3m.com/3M/en_IN/company-in/search/?Ntt=',
    'Livevi_VN':'https://www.3m.com.vn/3M/vi_VN/company-vn/search/?Ntt=', 'Prevvi_VN':'https://fuzeauth.3m.com/3M/vi_VN/company-vn/search/?Ntt=',
    'Livezh_CN':'https://www.3m.com.cn/3M/zh_CN/company-cn/search/?Ntt=', 'Prevzh_CN':'https://fuzeauth.3m.com/3M/zh_CN/company-cn/search/?Ntt=',
    'Liveen_ID':'https://www.3m.co.id/3M/en_ID/company-id/search/?Ntt=', 'Preven_ID':'https://fuzeauth.3m.com/3M/en_ID/company-id/search/?Ntt=',
    'Livezh_HK':'https://www.3m.com.hk/3M/zh_HK/company-hk/search/?Ntt=', 'Prevzh_HK':'https://fuzeauth.3m.com/3M/3M/zh_HK/company-hk/search/?Ntt=',
    'Liveen_HK':'https://www.3m.com.hk/3M/en_HK/company-hk/search/?Ntt=', 'Preven_HK':'https://fuzeauth.3m.com/3M/en_HK/company-hk/search/?Ntt=',
    'Livezh_TW':'https://www.3m.com.tw/3M/zh_TW/company-tw/search/?Ntt=', 'Prevzh_TW':'https://fuzeauth.3m.com/3M/zh_TW/company-tw/search/?Ntt='}

    wb = load_workbook("QA report.xlsx", data_only=True)
    ws0 = wb["Input"]
    ws1 = wb["PDPs"]

    key = ws0.cell(row=1, column=1).value + ws0.cell(row=1, column=2).value
    url = inputs[key] + str(ws1.cell(row=2, column=1).value)

    option = webdriver.ChromeOptions()
    option.headless = True
    browser = webdriver.Chrome(options=option)
    browser.maximize_window()
    browser.get(url)
    
    try:
        PDP = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "MMM--serpLink.js-ellipses")))
        ws1.cell(row=2, column=3).value = PDP.get_attribute("href")
        
    except:
        ws1.cell(row=2, column=3).value = "Not Found"

    finally:
        wb.save("QA report.xlsx")

    for x in range(3, ws1.max_row + 1):
        ID = ws1.cell(row=x, column=1).value
        NoURL = ws1.cell(row=x, column=3).value

        if ID == None:
            break    
        if NoURL == "Not Found" or NoURL == None:
            search = browser.find_element_by_id("js-searchBar")
            search.send_keys(ID)
            search.send_keys(Keys.ENTER)
            
            try:
                PDP = WebDriverWait(browser, 5).until(EC.presence_of_element_located((By.CLASS_NAME, "MMM--serpLink.js-ellipses")))
                ws1.cell(row=x, column=3).value = PDP.get_attribute("href") 

                progress = (x-1) / (ws1.max_row-1) * 100

                global p_var
                global pb
                p_var.set(progress)
                pb.update()

            except:
                ws1.cell(row=x, column=3).value = "Not Found"                     

            finally:
                wb.save("QA report.xlsx")
        
    browser.quit()

    for y in range(2, ws1.max_row + 1):
        ws1.cell(row=y, column=4).value = '=IFERROR(RIGHT(C{},LEN(C{})-FIND("○",SUBSTITUTE(C{},"/","○",LEN(C{})-LEN(SUBSTITUTE(C{},"/",""))-1))),"")'.format(y,y,y,y,y)

    for z in range(2, ws1.max_row + 1):
        ws1.cell(row=z, column=5).value = '=IF(A{}="","",IF(C{}="Not Found",C{},IF(LEFT(D{},1)="b",SUBSTITUTE(C{},"/d/","/dc/"),C{})))'.format(z,z,z,z,z,z)    
    
    wb.save("QA report.xlsx")
    os.startfile("QA report.xlsx")

def scrpt2():
    wb = load_workbook("QA report.xlsx", data_only=True)
    ws1 = wb["PDPs"]

    for i in range(2, ws1.max_row + 1):
        ws1.cell(row=i, column=6).value = '=IF(ISNUMBER(SEARCH("/dc/b",E{})),SUBSTITUTE(E{},D{},B{}&"/"),E{})'.format(i,i,i,i,i)

    for j in range(2, ws1.max_row + 1):
        ws1.cell(row=j, column=7).value = '=IF(A{}="","",IF(F{}="Not Found",F{},SUBSTITUTE(F{},"/V","/v")))'.format(j,j,j,j)
                                            
    wb.save("QA report.xlsx")
    os.startfile("QA report.xlsx")

def scrpt3():
    wb = load_workbook("QA report.xlsx", data_only=True)
    ws1 = wb["PDPs"]
    ws2 = wb["Product Details"]

    for i in range(2, ws1.max_row + 1):
        ws2.cell(row=i, column=1).value = ws1.cell(row=i, column=1).value 
        ws2.cell(row=i, column=2).value = ws1.cell(row=i, column=3).value 
        ws2.cell(row=i, column=3).value = ws1.cell(row=i, column=7).value 

    for x in range(2, ws2.max_row + 1):
        try:
            url = ws2.cell(row=x, column=3).value
            Product = ws2.cell(row=x, column=4).value      
            if Product == None:
                response = requests.get(url)
                response.raise_for_status()
                soup = BeautifulSoup(response.text, "lxml")
            
                MPFN = soup.find("h1", attrs={"itemprop":"name"}).get_text().strip()
                ws2.cell(row=x, column=4, value=MPFN) # Marketplace Formal Name

                Celum = soup.find("img", attrs={"data-type":"Main Picture"})
                ws2.cell(row=x, column=5).value = int(Celum["data-celum"])
                ws2.cell(row=x, column=6).value = Celum["src"]

                descriptions = soup.find_all("p", attrs={"class":"SNAPS--pdpTabDes"}) # Marketplace Description
                ws2.cell(row=x, column=7).value = descriptions[0].text
                
                if len(descriptions) == 3: # Marketplace Description Extended
                    extd1 = str(soup.find("div", attrs={"class":"MMM--grids-col MMM--grids-col_pdpMain"}))
                    extd2 = extd1.rsplit('<p class="SNAPS--pdpTabDes"></p>', 1)
                    extd3 = extd2[0].split('</p>', 1)
                    extd4 = extd3[1].strip()
                    extd5 = re.sub('<p class="SNAPS--pdpTabDes">',"", extd4, 1)
                    ws2.cell(row=x, column=8).value = re.sub('</p>',"", extd5, 1)
                    # ws2.cell(row=x, column=8).value = extd3[1].strip()
                    
                Bullets = [] # list to have all bullets
                bulletss = soup.find_all("ul", attrs={"class":"MMM--pdpList SNAPS--bullets"})
                if len(bulletss) >= 2:
                    bullets = bulletss[len(bulletss) - 1].find_all("li")
                else:
                    bullets = bulletss[0].find_all("li")
        
                for bullet in bullets:
                    Bullets.append(bullet.get_text().strip())
            
                for y in range(len(Bullets)):
                    ws2.cell(row=x, column=y+9, value=Bullets[y])
                
            progress = (x-1) / (ws1.max_row-1) * 100
            global p_var
            global pb
            p_var.set(progress)
            pb.update()

        except:
            pass
        
        for row in ws2.rows:
            for cell in row:
                cell.alignment = Alignment(horizontal="center",wrap_text=True)
                cell.border = Border(left=Side(style="thin", color="A9A9A9"), right=Side(style="thin", color="A9A9A9"),top=Side(style="thin", color="A9A9A9"),bottom=Side(style="thin", color="A9A9A9") )
                cell.font = Font(name="Arial", size=9)  
                    
        for row2 in range(2, ws2.max_row + 1):
            ws2.row_dimensions[row2].height = 24

        Report_Headers = ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1", "I1", "J1", "K1", "L1", "M1", "N1", "O1"]
        for Header in Report_Headers:
            ws2[Header].font = Font(color="FFFAF0", name="Arial", size=9, bold=True)
            ws2[Header].fill = PatternFill(fgColor="104E8B", fill_type="solid")

    wb.save("QA report.xlsx")
    os.startfile("QA report.xlsx")

def scrpt4():
    wb = load_workbook("QA report.xlsx", data_only=True)
    ws1 = wb["PDPs"]
    ws4 = wb["Specifications"]

    for i in range(2, ws1.max_row + 1):
        ws4.cell(row=i, column=1).value = ws1.cell(row=i, column=1).value 
        ws4.cell(row=i, column=2).value = ws1.cell(row=i, column=7).value 

    for x in range(2, ws4.max_row + 1):
        try:
            url = ws4.cell(row=x, column=2).value
            Product = ws4.cell(row=x, column=3).value
            if Product == None:
                response = requests.get(url)
                response.raise_for_status()
                soup = BeautifulSoup(response.text, "lxml")

                MPFN = soup.find("h1", attrs={"itemprop":"name"}).get_text().strip()
                ws4.cell(row=x, column=3, value=MPFN) # Marketplace Formal Name

                specs = soup.find_all("tr", attrs={"class":"MMM--dat-row"})
                ATT = []
                VAL = []

                for j in range(len(specs)):
                    try:
                        ATT.append(specs[j].find("div", attrs={"class":"MMM--dat-cell SNAPS-specAdjust"}).get_text(strip=True))
                        VAL.append(specs[j].find("div", attrs={"class":"MMM--dat-cell dat-cell_specDefinition SNAPS-specAdjust SNAPS--SpecTbl-RowCols SNAPS--SpecTbl-Overflow"}).get_text(strip=True))
                    except:
                        pass

                for k in range(len(ATT)):
                    ws4.cell(row=x, column=k+4).value = ATT[k] +" : " + VAL[k] 
        
            progress = (x-1) / (ws1.max_row-1) * 100
            global p_var
            global pb
            p_var.set(progress)
            pb.update()

        except:
            pass
            
    for row in ws4.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal="center",wrap_text=True)
            cell.border = Border(left=Side(style="thin", color="A9A9A9"), right=Side(style="thin", color="A9A9A9"),top=Side(style="thin", color="A9A9A9"),bottom=Side(style="thin", color="A9A9A9") )
            cell.font = Font(name="Arial", size=9)  
        
    for row2 in range(2, ws4.max_row + 1):
        ws4.row_dimensions[row2].height = 24

    for cell in ws4[1]:
        cell.font = Font(color="FFFAF0", name="Arial", size=9, bold=True)
        cell.fill = PatternFill(fgColor="104E8B", fill_type="solid")

    wb.save("QA report.xlsx")  
    os.startfile("QA report.xlsx")

def scrpt5():
    wb = load_workbook("QA report.xlsx", data_only=True)
    ws1 = wb["PDPs"]
    ws3 = wb["WTBs"]

    for i in range(2, ws1.max_row + 1):
        ws3.cell(row=i, column=1).value = ws1.cell(row=i, column=1).value 
        ws3.cell(row=i, column=2).value = ws1.cell(row=i, column=7).value 

    for x in range(2, ws3.max_row + 1):
        try:
            url = ws3.cell(row=x, column=2).value
            Product = ws3.cell(row=x, column=3).value
                
            if Product == None:
                response = requests.get(url)
                response.raise_for_status()
                soup = BeautifulSoup(response.text, "lxml")

                MPFN = soup.find("h1", attrs={"itemprop":"name"}).get_text().strip()
                ws3.cell(row=x, column=3, value=MPFN) # Marketplace Formal Name

                try:
                    wtbs = soup.find("div", attrs={"class":"SNAPS--wtb_section SNAPS--wtb_section_original"}).find_all("wtb")
                    Labels = [] # list to have all WTB labels
                    URLs = []
                    for wtb in wtbs:
                        Labels.append(wtb["label"])
                        URLs.append(wtb["url"])
                    for j in range(len(Labels)):
                        ws3.cell(row=x, column=j+4, value = Labels[j] + " : " + URLs[j])

                except:
                    ws3.cell(row=x, column=4, value="NO WTB")

            progress = (x-1) / (ws1.max_row-1) * 100
            global p_var
            global pb
            p_var.set(progress)
            pb.update()
        
        except:
            pass

    for row in ws3.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal="center",wrap_text=True)
            cell.border = Border(left=Side(style="thin", color="A9A9A9"), right=Side(style="thin", color="A9A9A9"),top=Side(style="thin", color="A9A9A9"),bottom=Side(style="thin", color="A9A9A9") )
            cell.font = Font(name="Arial", size=9)  
        
    for row2 in range(2, ws3.max_row + 1):
        ws3.row_dimensions[row2].height = 24

    Report_Headers = ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1"]
    for Header in Report_Headers:
        ws3[Header].font = Font(color="FFFAF0", name="Arial", size=9, bold=True)
        ws3[Header].fill = PatternFill(fgColor="104E8B", fill_type="solid")

    wb.save("QA report.xlsx")  
    os.startfile("QA report.xlsx")

def script8():
    wb = load_workbook("QA report.xlsx", data_only=True)
    ws1 = wb["PDPs"]
    ws3 = wb["WTBs"]

    for i in range(2, ws1.max_row + 1):
        ws3.cell(row=i, column=1).value = ws1.cell(row=i, column=1).value 
        ws3.cell(row=i, column=2).value = ws1.cell(row=i, column=7).value 

    for x in range(2, ws3.max_row + 1):
        try:
            url = ws3.cell(row=x, column=2).value
            Product = ws3.cell(row=x, column=3).value
                
            if Product == None:
                response = requests.get(url)
                response.raise_for_status()
                soup = BeautifulSoup(response.text, "lxml")

                MPFN = soup.find("h1", attrs={"itemprop":"name"}).get_text().strip()
                ws3.cell(row=x, column=3, value=MPFN) # Marketplace Formal Name

                try:
                    wtbs = soup.find("div", attrs={"class":"MMM--selectionBox--cntnr"}).find_all("div")
                    Subs = [] # list to have all WTB labels
                    Temps = []
                    for wtb in wtbs:
                        Subs.append(wtb["data-subid"])
                        Temps.append(wtb["data-widgetid"])
                    for j in range(len(Subs)):
                        ws3.cell(row=x, column=j+4, value = Subs[j] + " : " + Temps[j])
                except:
                    ws3.cell(row=x, column=4, value="NO WTB")

            progress = (x-1) / (ws1.max_row-1) * 100
            global p_var
            global pb
            p_var.set(progress)
            pb.update()
        
        except:
            pass

    for row in ws3.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal="center",wrap_text=True)
            cell.border = Border(left=Side(style="thin", color="A9A9A9"), right=Side(style="thin", color="A9A9A9"),top=Side(style="thin", color="A9A9A9"),bottom=Side(style="thin", color="A9A9A9") )
            cell.font = Font(name="Arial", size=9)  
       
    for row2 in range(2, ws3.max_row + 1):
        ws3.row_dimensions[row2].height = 24

    Report_Headers = ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1"]
    for Header in Report_Headers:
        ws3[Header].font = Font(color="FFFAF0", name="Arial", size=9, bold=True)
        ws3[Header].fill = PatternFill(fgColor="104E8B", fill_type="solid")

    wb.save("QA report.xlsx")  
    os.startfile("QA report.xlsx")

def scrpt6():
    wb = load_workbook("QA report.xlsx", data_only=True)
    ws0 = wb["Input"]
    ws2 = wb["Product Details"]
    ws3 = wb["Original"]

    locale = ws0.cell(row=1, column=2).value
    row1 = ["Marketplace Formal Name","Main Picture", "Marketplace Description", "Marketplace Description Extended", "Bullet 1", "Bullet 2", "Bullet 3", "Bullet 4", "Bullet 5", "Bullet 6", "Bullet 7"]

    col1 = []
    col2 = [] 
    for x in range(2, ws3.max_column + 1): 
        if ws3.cell(row=3, column=x).value == locale: # target locale only
            if ws3.cell(row=2, column=x).value in row1:
                col1.append(ws3.cell(row=2, column=x).value)
                col2.append(ws3.cell(row=2, column=x).column)

    col3 = dict(zip(col1,col2)) # Attributes & Column number for the target locale in the Original template

    for MPFN in range(2, ws2.max_row + 1): # Color the different values from those in the Original
        try:
            if ws2.cell(row=MPFN, column=4).value != ws3.cell(row = MPFN + 2, column = col3["Marketplace Formal Name"]).value:
                ws2.cell(row=MPFN, column=4).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass

    for MP in range(2, ws2.max_row + 1):
        try:
            if ws2.cell(row=MP, column=5).value != ws3.cell(row = MP + 2, column = col3["Main Picture"]).value:
                ws2.cell(row=MP, column=5).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass
                    
    for Dsc in range(2, ws2.max_row + 1):
        try:
            if ws2.cell(row=Dsc, column=7).value != ws3.cell(row = Dsc + 2, column = col3["Marketplace Description"]).value:
                ws2.cell(row=Dsc, column=7).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass
                    
    for Extd in range(2, ws2.max_row + 1):
        try:
            if ws2.cell(row=Extd, column=8).value != ws3.cell(row = Extd + 2, column = col3["Marketplace Description Extended"]).value:
                ws2.cell(row=Extd, column=8).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass
                    
    for B1 in range(2, ws2.max_row + 1):
        try:
            if ws2.cell(row=B1, column=9).value != ws3.cell(row = B1 + 2, column = col3["Bullet 1"]).value:
                ws2.cell(row=B1, column=9).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass
                    
    for B2 in range(2, ws2.max_row + 1):
        try:
            if ws2.cell(row=B2, column=10).value != ws3.cell(row = B2 + 2, column = col3["Bullet 2"]).value:
                ws2.cell(row=B2, column=10).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass
                    
    for B3 in range(2, ws2.max_row + 1):
        try:
            if ws2.cell(row=B3, column=11).value != ws3.cell(row = B3 + 2, column = col3["Bullet 3"]).value:
                ws2.cell(row=B3, column=11).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass
                    
    for B4 in range(2, ws2.max_row + 1):
        try:
            if ws2.cell(row=B4, column=12).value != ws3.cell(row = B4 + 2, column = col3["Bullet 4"]).value:
                ws2.cell(row=B4, column=12).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass
                    
    for B5 in range(2, ws2.max_row + 1):
        try:
            if ws2.cell(row=B5, column=13).value != ws3.cell(row = B5 + 2, column = col3["Bullet 5"]).value:
                ws2.cell(row=B5, column=13).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass
                    
    for B6 in range(2, ws2.max_row + 1):
        try:
            if ws2.cell(row=B6, column=14).value != ws3.cell(row = B6 + 2, column = col3["Bullet 6"]).value:
                ws2.cell(row=B6, column=14).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass
                    
    for B7 in range(2, ws2.max_row + 1):
        try:
            if ws2.cell(row=B7, column=15).value != ws3.cell(row = B7 + 2, column = col3["Bullet 7"]).value:
                ws2.cell(row=B7, column=15).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass
                    
    for y in ws2[1]: # Columns which Values from the Original are to be pasted in.
        if y.value in col1:
            ws2.insert_cols(y.column)
        
    for i in range(4, ws2.max_column +1):
        if ws2.cell(row=1, column=i).value == None:
            ws2.cell(row=1, column=i).value = "Original"
            ws2.cell(row=1, column=i).alignment = Alignment(horizontal="center")
            ws2.cell(row=1, column=i).font = Font(color="FFFAF0", name="Arial", size=9, bold=True)
            ws2.cell(row=1, column=i).fill = PatternFill(fgColor="CD5555", fill_type="solid")

            for j in range(2, ws2.max_row + 1):
                try:
                    ws2.cell(row=j, column=i).value = ws3.cell(row = j + 2, column = col3[ws2.cell(row=1, column = i + 1).value]).value
                    ws2.cell(row=j, column=i).alignment = Alignment(horizontal="center", wrap_text=True)
                    ws2.cell(row=j, column=i).font = Font(name="Arial", size=9)
                    ws2.cell(row=j, column=i).border = Border(left=Side(style="thin", color="A9A9A9"), right=Side(style="thin", color="A9A9A9"),top=Side(style="thin", color="A9A9A9"),bottom=Side(style="thin", color="A9A9A9") )
                except:
                    pass

    wb.save("QA report.xlsx")
    os.startfile("QA report.xlsx")

def scrpt7():
    wb = load_workbook("QA report.xlsx", data_only=True)
    ws0 = wb["Input"]
    ws3 = wb["Original"]
    ws5 = wb["PISA"]

    locale = ws0.cell(row=1, column=2).value
    row1 = ["Marketplace Formal Name","Main Picture", "Marketplace Description", "Marketplace Description Extended", "Bullet 1", "Bullet 2", "Bullet 3", "Bullet 4", "Bullet 5", "Bullet 6", "Bullet 7", "FUZE Market Level 1", "FUZE Market Level 2"]
    # Set of Attributes whose data is to be complated between Original and PISA tabs

    col1 = [] # Atrbitues
    col2 = [] # Column numbers
    for x in range(2, ws3.max_column + 1): 
        if ws3.cell(row=3, column=x).value == locale:
            if ws3.cell(row=2, column=x).value in row1:
                col1.append(ws3.cell(row=2, column=x).value)
                col2.append(ws3.cell(row=2, column=x).column)

    col3 = dict(zip(col1,col2)) # Set of Attributes & Column numbers for the Target Locale in Original tab

    col4 = [] # Atrbitues for the Target Locale in Pisa tab
    col5 = [] # Column number
    for y in range(2, ws5.max_column + 1): # Before moving columns of comparable to the left, identity Atttibutes
        if ws5.cell(row=3, column=y).value == locale:
            if ws5.cell(row=2, column=y).value in row1:
                col4.append(ws5.cell(row=2, column=y).value)
                    
    ws5.insert_cols(7, len(col4)) # Adding(inserting) columns for data to be copied from columns of comparable in PISA tab 

    for z in range(2, ws5.max_column + 1): # After adding columns, identity Column numbers
        if ws5.cell(row=3, column=z).value == locale:
            if ws5.cell(row=2, column=z).value in row1:
                col5.append(ws5.cell(row=2, column=z).column)

    for idx1, p in enumerate(col5): # Move Columns of comparable data in PISA tab
        for q in ws5[get_column_letter(p)]:
            move_cell = q.offset(column = 7 + idx1 - p)
            move_cell.value = q.value
            move_cell.alignment = Alignment(horizontal="center", wrap_text=True)
            move_cell.font = Font(name="Arial", size=9)

    for idx2, r in enumerate(col5):  # After moving, Delete old columns of comparable data in PISA tab
        ws5.delete_cols(r-idx2)

    col5.clear()
    for s in range(2, ws5.max_column + 1):
        if ws5.cell(row=3, column=s).value == locale:
            if ws5.cell(row=2, column=s).value in row1:
                col5.append(ws5.cell(row=2, column=s).column)

    col6 = dict(zip(col4,col5)) # Set of Attributes & column numbers for the target locale in PISA tab

    for MPFN in range(4, ws5.max_row + 1): # Marektplace Formal Name, Color the Value in PISA tab which is not same as that in Original tab
        try:
            if ws5.cell(row=MPFN, column = col6["Marketplace Formal Name"]).value != ws3.cell(row=MPFN, column = col3["Marketplace Formal Name"]).value:
                ws5.cell(row=MPFN, column = col6["Marketplace Formal Name"]).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass

    for Pic in range(4, ws5.max_row + 1):  # Main Picture
        try:
            if ws5.cell(row=Pic, column = col6["Main Picture"]).value != ws3.cell(row = Pic + 2, column = col3["Main Picture"]).value:
                ws5.cell(row=Pic, column = col6["Main Picture"]).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass

    for Dsc in range(4, ws5.max_row + 1): # Marketplace Description
        try:
            if ws5.cell(row=Dsc, column = col6["Marketplace Description"]).value != ws3.cell(row=Dsc, column = col3["Marketplace Description"]).value:
                ws5.cell(row=Dsc, column = col6["Marketplace Description"]).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass

    for Extd in range(4, ws5.max_row + 1): # Marketplace Description Extended
        try:
            if ws5.cell(row= Extd, column = col6["Marketplace Description Extended"]).value != ws3.cell(row=Extd, column = col3["Marketplace Description Extended"]).value:
                ws5.cell(row= Extd, column = col6["Marketplace Description Extended"]).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass

    for B1 in range(4, ws5.max_row + 1): # Bullet 1
        try:
            if ws5.cell(row=B1, column = col6["Bullet 1"]).value != ws3.cell(row=B1, column = col3["Bullet 1"]).value:
                ws5.cell(row=B1, column = col6["Bullet 1"]).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass

    for B2 in range(4, ws5.max_row + 1): # Bullet 2
        try:
            if ws5.cell(row=B2, column = col6["Bullet 2"]).value != ws3.cell(row=B2, column = col3["Bullet 2"]).value:
                ws5.cell(row=B2, column = col6["Bullet 2"]).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass

    for B3 in range(4, ws5.max_row + 1): # Bullet 3
        try:
            if ws5.cell(row=B3, column = col6["Bullet 3"]).value != ws3.cell(row=B3, column = col3["Bullet 3"]).value:
                ws5.cell(row=B3, column = col6["Bullet 3"]).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass

    for B4 in range(4, ws5.max_row + 1): # Bullet 4
        try:
            if ws5.cell(row=B4, column = col6["Bullet 4"]).value != ws3.cell(row=B4, column = col3["Bullet 4"]).value:
                ws5.cell(row=B4, column = col6["Bullet 4"]).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass

    for B5 in range(4, ws5.max_row + 1): # Bullet 5
        try:
            if ws5.cell(row=B5, column = col6["Bullet 5"]).value != ws3.cell(row=B5, column = col3["Bullet 5"]).value:
                ws5.cell(row=B5, column = col6["Bullet 5"]).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass

    for B6 in range(4, ws5.max_row + 1): # Bullet 6
        try:
            if ws5.cell(row=B6, column = col6["Bullet 6"]).value != ws3.cell(row=B6, column = col3["Bullet 6"]).value:
                ws5.cell(row=B6, column = col6["Bullet 6"]).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass

    for B7 in range(4, ws5.max_row + 1): # Bullet 7
        try:
            if ws5.cell(row=B7, column = col6["Bullet 7"]).value != ws3.cell(row=B7, column = col3["Bullet 7"]).value:
                ws5.cell(row=B7, column = col6["Bullet 7"]).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass    

    for MKT1 in range(4, ws5.max_row + 1): # FUZE Market Level 1
        try:
            if ws5.cell(row=MKT1, column = col6["FUZE Market Level 1"]).value != ws3.cell(row=MKT1, column = col3["FUZE Market Level 1"]).value:
                ws5.cell(row=MKT1, column = col6["FUZE Market Level 1"]).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass

    for MKT2 in range(4, ws5.max_row + 1): # FUZE Market Level 2
        try:
            if ws5.cell(row=MKT2, column = col6["FUZE Market Level 2"]).value != ws3.cell(row=MKT2, column = col3["FUZE Market Level 2"]).value:
                ws5.cell(row=MKT2, column = col6["FUZE Market Level 2"]).fill = PatternFill(fgColor="FFEC8B", fill_type="solid")
        except:
            pass
 
    for Header in ws5[2:3]: # Header styles
        for cell in Header:
            if cell.column == 1:
                continue
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.font = Font(color="FFFAF0", name="Arial", size=9, bold=True)
            cell.fill = PatternFill(fgColor="104E8B", fill_type="solid")

    for h in ws5[2]: # Columns for data to be copied from the Original tab
        if h.value in row1:
            if h.offset(row=1).value ==locale:
                ws5.insert_cols(h.column)
                h.offset(column=-1).value = "Original"
                h.offset(column=-1).font = Font(color="FFFAF0", name="Arial", size=9, bold=True)
                h.offset(column=-1).fill = PatternFill(fgColor="CD5555", fill_type="solid")
                h.offset(row=1, column=-1).fill = PatternFill(fgColor="CD5555", fill_type="solid")

    col5.clear()
    for i in range(2, ws5.max_column + 1): # Column numbers for the target locale in PISA tab
        if ws5.cell(row=3, column=i).value == locale:
            if ws5.cell(row=2, column=i).value in row1:
                col5.append(ws5.cell(row=2, column=i).column)

    for j in col5: # Paste data from Original tab to Original columns in PISA tab
        for k in range(4, ws5.max_row + 1):
            try:
                ws5.cell(row=k, column = j - 1).value = ws3.cell(row=k, column = col3[ws5.cell(row=2, column = j).value]).value
                ws5.cell(row=k, column = j - 1).alignment = Alignment(horizontal="center", wrap_text=True)
                ws5.cell(row=k, column = j - 1).font = Font(name="Arial", size=9)
            except:
                pass

    wb.save("QA report.xlsx")
    os.startfile("QA report.xlsx")

myFont2 = font.Font(family='Franklin Gothic Medium', size=10, weight='bold')
btn1 = Button(root, text="PDP (Search)", font=myFont2, bg='powder blue', padx=5, pady=5, command=scrpt1)
btn2 = Button(root, text="PDP (Create)", font=myFont2, bg='white', padx=5, pady=5, command=scrpt2)
btn3 = Button(root, text="Product Details", font=myFont2, bg='white', padx=5, pady=5, command=scrpt3)
btn4 = Button(root, text="SPecifications", font=myFont2, bg='white', padx=5, pady=5, command=scrpt4)
btn5 = Button(root, text="WTB/CTA", font=myFont2, bg='white', padx=5, pady=5, command=scrpt5)
btn8 = Button(root, text="WTB (New)", font=myFont2, bg='white', padx=5, pady=5, command=script8)
btn6 = Button(root, text="Data Comparison (Web)", font=myFont2, bg='white', padx=5, pady=5, command=scrpt6)
btn7 = Button(root, text="Data comparison (PISA)", font=myFont2, bg='white', padx=5, pady=5, command=scrpt7)
label1 = Label(root, text="[PROGRESS]", font=myFont2)

btn1.grid(row=0, column=0, sticky=N+E+W+S, padx=1, pady=1)
btn2.grid(row=0, column=1, sticky=N+E+W+S, padx=1, pady=1)
btn3.grid(row=1, column=0, sticky=N+E+W+S, padx=1, pady=1)
btn4.grid(row=1, column=1, sticky=N+E+W+S, padx=1, pady=1)
btn5.grid(row=2, column=0, sticky=N+E+W+S, padx=1, pady=1)
btn8.grid(row=2, column=1, sticky=N+E+W+S, padx=1, pady=1)
btn6.grid(row=3, column=0, columnspan=2, sticky=N+E+W+S, padx=1, pady=1)
btn7.grid(row=4, column=0, columnspan=2, sticky=N+E+W+S, padx=1, pady=1)
label1.grid(row=5, column=0, rowspan=2, columnspan=2, sticky=E+W+S)

root.mainloop()